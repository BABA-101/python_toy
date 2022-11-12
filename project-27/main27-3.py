# 엑셀에서 학교명과 주소 읽어 x,y 좌표로 변경
# 변경된 값을 새로운 엑셀 파일로 만들기

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

path=r'project-27\고등교육기관 하반기 주소록(2020).xlsx'
df_from_excel = pd.read_excel(path, engine='openpyxl')
df_from_excel.columns=df_from_excel.loc[4].tolist() 
df_from_excel=df_from_excel.drop(index=list(range(0,5))) #제거

url='http://api.vworld.kr/req/address?'
params='service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type='ROAD' #도로명
road_type2='PARCEL' #지번
address='&address='
keys='&key='
primary_key='75AE8ECD-3956-3C76-82EF-F9FE6B9082CA'

def request_geo(road):
    page=requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data=page.json()
    if json_data['response']['status']=='OK':
        x=json_data['response']['result']['point']['x']
        y=json_data['response']['result']['point']['y']
        return x,y
    else:
        x=0
        y=0
        return x,y

try:
    # data_only: True면 수식 아닌 값으로 불러옴. 로드할건데 없으면 예외
    wb=load_workbook(r'project-27\address_xy.xlsx', data_only=True)
    sheet=wb.active
except:
    wb=Workbook()
    sheet=wb.active

university_list=df_from_excel['학교명'].tolist()
address_list=df_from_excel['주소'].tolist()


for num,value in enumerate(address_list):
    # sub(pattern, repl, string): string에서 pattern에 맞는 부분을 repl로 대체
    # value에서 괄호를 제거하기 위해 정규표현식 작성
    addr = re.sub(r'[()]', '', value)
    # print(addr)
    x,y = request_geo(addr)
    sheet.append([university_list[num],addr,x,y])

wb.save(r"project-27\address_xy.xlsx")