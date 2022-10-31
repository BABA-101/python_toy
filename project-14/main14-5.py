#엑셀 파일에서 대량으로 이메일 보내는 코드 만들기
# A행: 보낼 이메일 주소
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb=load_workbook(r"project-14\email_list.xlsx",data_only=True)
load_ws=load_wb.active

for i in range(1,load_ws.max_row +1):
    recv_email_value=load_ws.cell(i,1).value
    print("성공: ",recv_email_value)
    try:
        # 메일주소 에러 없을 때 메일전송
        send_email=""
        send_pwd=""

        recv_email=""

        smtp_name="smtp.gmail.com"
        smtp_port=587

        msg=MIMEMultipart()
        
        msg['Subject']="엑셀에서 메일 주소 읽어 자동으로 보내는 메일이다"
        msg['From']=send_email
        msg['To']=recv_email

        text="""
            메일 내용
            ㄳㄳ
        """

        msg.attach(MIMEText(text))

        s=smtplib.SMTP(smtp_name,smtp_port)
        s.starttls()
        s.login(send_email,recv_email,msg.as_string())
        s.quit()
    except:
        # 이메일 주소의 에러가 있을 때 에러가 발생한 메일주소 출력
        print("에러: ",recv_email_value)