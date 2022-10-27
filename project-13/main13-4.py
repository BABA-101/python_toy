# 수집한 이메일 주소를 엑셀에 저장하는 코드
import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url='https://www.news1.kr/articles/?4846342&kakao_from=mainnews'

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

res=requests.get(url,headers=headers)

results=re.findall(r'[\w\.-]+@[\w\.-]+',res.text)
results=list(set(results))
print(results)

try:
    # email.xlsx 파일이 존재하여 로드할 수 있는경우, 파일을 읽는다.
    wb=load_workbook(r'project-13\email.xlsx',data_only=True)
    sheet=wb.active
except:
    # email.xlsx 파일이 없다면 새로운 엑셀 파일 생성
    wb=Workbook()
    sheet=wb.active

# 이메일을 수집한 만큼 반복
for result in results:
    sheet.append([result])

wb.save(r'project-13\email.xlsx')