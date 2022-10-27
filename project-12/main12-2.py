# 엑셀에 저장한 정보를 불러오는 코드
# 엑셀에서 파일을 읽기 위해 openpyxl 라이브러리에서 load_workbook를 불러옴
from openpyxl import load_workbook

load_wb=load_workbook(r'project-12\수료증명단.xlsx')
# 읽어온 엑셀 파일에서 활성화된 시트 불러오기
load_ws=load_wb.active

name_list=[]
birth_list=[]
ho_list=[]

for i in range(1,load_ws.max_row + 1):
    name_list.append(load_ws.cell(i,1).value)
    birth_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)

print(name_list)
print(birth_list)
print(ho_list)
